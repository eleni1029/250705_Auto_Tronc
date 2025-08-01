#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 路徑分析器 - 調整版
分析 Excel 課程結構檔案中的路徑，檢查檔案存在性並解析 HTML 內容
調整點：
1. 修改"路徑"欄位的值為：資源庫路徑/原路徑
2. 遺失明細移除資料夾路徑前綴
3. 系統路徑維持原邏輯：資料夾路徑/原路徑
"""

import os
import re
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Tuple, Set, Optional
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from urllib.parse import urljoin, urlparse
import html
from html.parser import HTMLParser


class MediaFileParser(HTMLParser):
    """HTML 媒體檔案解析器"""
    
    def __init__(self):
        super().__init__()
        self.media_files = set()
        self.current_tag = None
        
        # 支援的媒體檔案格式
        self.video_extensions = {
            'mpg', 'mpeg', 'mp4', 'mkv', 'avi', 'mov', 'wmv', 'flv', 'webm', 'm3u8'
        }
        
        self.audio_extensions = {
            'mp3', 'wav', 'aac', 'ogg', 'flac', 'wma', 'midi', 'mid'
        }
        
        self.image_extensions = {
            'jpg', 'jpeg', 'png', 'gif', 'bmp', 'svg', 'webp', 'tiff', 'ico'
        }
        
        self.pdf_extensions = {'pdf'}
        
        self.all_extensions = (
            self.video_extensions | self.audio_extensions | 
            self.image_extensions | self.pdf_extensions
        )
    
    def handle_starttag(self, tag, attrs):
        """處理開始標籤"""
        self.current_tag = tag
        attrs_dict = dict(attrs)
        
        # 根據不同標籤提取媒體檔案路徑
        if tag == 'img':
            # <img src="..." data-src="...">
            for attr in ['src', 'data-src', 'data-original']:
                if attr in attrs_dict:
                    self._check_and_add_media_file(attrs_dict[attr])
        
        elif tag == 'video':
            # <video src="..." poster="...">
            for attr in ['src', 'poster']:
                if attr in attrs_dict:
                    self._check_and_add_media_file(attrs_dict[attr])
        
        elif tag == 'audio':
            # <audio src="...">
            if 'src' in attrs_dict:
                self._check_and_add_media_file(attrs_dict['src'])
        
        elif tag == 'source':
            # <source src="...">
            if 'src' in attrs_dict:
                self._check_and_add_media_file(attrs_dict['src'])
        
        elif tag == 'a':
            # <a href="file.pdf">
            if 'href' in attrs_dict:
                self._check_and_add_media_file(attrs_dict['href'])
        
        elif tag == 'embed':
            # <embed src="...">
            if 'src' in attrs_dict:
                self._check_and_add_media_file(attrs_dict['src'])
        
        elif tag == 'object':
            # <object data="...">
            if 'data' in attrs_dict:
                self._check_and_add_media_file(attrs_dict['data'])
        
        elif tag == 'iframe':
            # <iframe src="...">
            if 'src' in attrs_dict:
                self._check_and_add_media_file(attrs_dict['src'])
        
        elif tag == 'track':
            # <track src="...">
            if 'src' in attrs_dict:
                self._check_and_add_media_file(attrs_dict['src'])
    
    def _check_and_add_media_file(self, file_path: str):
        """檢查並添加媒體檔案"""
        if not file_path:
            return
        
        # 清理路徑
        file_path = file_path.strip()
        
        # 跳過外部連結和特殊協議
        if file_path.startswith(('http://', 'https://', 'ftp://', 'mailto:', 'javascript:', 'data:')):
            return
        
        # 跳過錨點連結
        if file_path.startswith('#'):
            return
        
        # 移除查詢參數和錨點
        clean_path = file_path.split('?')[0].split('#')[0]
        
        # 檢查檔案副檔名
        if clean_path:
            extension = Path(clean_path).suffix.lower().lstrip('.')
            if extension in self.all_extensions:
                self.media_files.add(file_path)
    
    def get_media_files(self) -> Set[str]:
        """獲取找到的媒體檔案"""
        return self.media_files


class ExcelPathAnalyzer:
    """Excel 路徑分析器類別"""
    
    def __init__(self, output_dir: str = "5_to_be_executed"):
        """
        初始化路徑分析器
        
        Args:
            output_dir: 輸出目錄路徑
        """
        self.output_dir = Path(output_dir)
        self.selected_excel_file = None
        self.output_file = None
        self.timestamp = None
        
        self.stats = {
            'sheets_processed': 0,
            'paths_analyzed': 0,
            'paths_modified': 0,
            'html_files_parsed': 0,
            'errors': 0
        }
        
        # 設定日誌
        self._setup_logging()
    
    def _setup_logging(self):
        """設定日誌系統"""
        log_filename = f"excel_path_analyzer_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        log_dir = Path("../log")
        log_dir.mkdir(parents=True, exist_ok=True)
        log_path = log_dir / log_filename

        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_path, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )

        self.logger = logging.getLogger(__name__)
        self.logger.info("=== 開始 Excel 路徑分析作業 ===")
        self.logger.info(f"輸出目錄: {self.output_dir.absolute()}")
        self.logger.info(f"Log 檔案儲存於: {log_path}")
    
    def _find_excel_files(self) -> List[Path]:
        """找到所有 course_structures 開頭的 Excel 檔案（不包含 report 和 analyzed）"""
        if not self.output_dir.exists():
            return []
        
        pattern = "course_structures_*.xlsx"
        excel_files = []
        
        for file_path in self.output_dir.glob(pattern):
            # 排除包含 report 或 analyzed 的檔案
            if "report" not in file_path.name.lower() and "analyzed" not in file_path.name.lower():
                excel_files.append(file_path)
        
        return sorted(excel_files, key=lambda x: x.stat().st_mtime, reverse=True)
    
    def _select_excel_file(self) -> bool:
        """讓用戶選擇要分析的 Excel 檔案"""
        excel_files = self._find_excel_files()
        
        if not excel_files:
            print("❌ 在 5_to_be_executed 目錄中未找到可分析的 course_structures 開頭的 Excel 檔案")
            return False
        
        print(f"\n找到 {len(excel_files)} 個可分析的 Excel 檔案：")
        print("=" * 50)
        
        for i, file_path in enumerate(excel_files, 1):
            file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
            file_size = file_path.stat().st_size
            print(f"{i}. {file_path.name}")
            print(f"   修改時間: {file_time.strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"   檔案大小: {file_size:,} bytes")
            print()
        
        # 用戶選擇
        while True:
            try:
                print(f"請選擇要分析的 Excel 檔案 (1-{len(excel_files)}) [輸入 '0' 使用預設: 1]: ", end="", flush=True)
                choice = input().strip()
                if not choice:
                    print("⚠️ 請輸入有效值，或輸入 '0' 使用預設值")
                    continue
                if choice == '0':
                    choice = "1"
                
                index = int(choice) - 1
                if 0 <= index < len(excel_files):
                    self.selected_excel_file = excel_files[index]
                    print(f"✅ 已選擇: {self.selected_excel_file.name}")
                    return True
                else:
                    print(f"❌ 請輸入 1 到 {len(excel_files)} 之間的數字")
            except ValueError:
                print("❌ 請輸入有效的數字")
            except KeyboardInterrupt:
                print("\n❌ 操作已取消")
                return False
    
    def _get_sheet_info(self, worksheet) -> Dict[str, str]:
        """從工作表獲取資源庫路徑和資料夾路徑（固定從A2和A3讀取）"""
        try:
            info = {'resource_path': '', 'folder_path': ''}
            
            # 從A2讀取資源庫路徑
            a2_cell = worksheet.cell(row=2, column=1).value
            if a2_cell:
                a2_text = str(a2_cell).strip()
                if '資源庫路徑：' in a2_text:
                    info['resource_path'] = a2_text.replace('資源庫路徑：', '').strip()
                    self.logger.debug(f"從A2找到資源庫路徑: {info['resource_path']}")
            
            # 從A3讀取資料夾路徑
            a3_cell = worksheet.cell(row=3, column=1).value
            if a3_cell:
                a3_text = str(a3_cell).strip()
                if '資料夾路徑：' in a3_text:
                    info['folder_path'] = a3_text.replace('資料夾路徑：', '').strip()
                    self.logger.debug(f"從A3找到資料夾路徑: {info['folder_path']}")
            
            return info
            
        except Exception as e:
            self.logger.error(f"獲取工作表資訊失敗: {e}")
            return {'resource_path': '', 'folder_path': ''}
    
    def _find_column_positions(self, worksheet) -> Dict[str, int]:
        """動態找到欄位位置 - 搜尋整個工作表"""
        positions = {}
        
        # 搜尋整個工作表前20行，找到所有相關欄位
        for row in range(1, min(worksheet.max_row + 1, 21)):  # 搜尋前20行
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    cell_text = str(cell_value).strip()
                    
                    # 記錄找到的欄位位置和所在行
                    if cell_text == "學習活動" and 'learning_activity' not in positions:
                        positions['learning_activity'] = {'col': col, 'row': row}
                    elif cell_text == "類型" and 'type' not in positions:
                        positions['type'] = {'col': col, 'row': row}
                    elif cell_text == "路徑" and 'path' not in positions:
                        positions['path'] = {'col': col, 'row': row}
        
        self.logger.info(f"找到欄位位置: {positions}")
        return positions
    
    def _find_data_start_row(self, worksheet, positions: Dict[str, int]) -> int:
        """動態找到資料開始的行號"""
        # 找到欄位標題的最大行號，資料從下一行開始
        max_header_row = 0
        for field_info in positions.values():
            if isinstance(field_info, dict) and 'row' in field_info:
                max_header_row = max(max_header_row, field_info['row'])
        
        data_start_row = max_header_row + 1
        self.logger.info(f"資料開始行號: {data_start_row}")
        return data_start_row
    
    def _is_external_link(self, file_path: str) -> bool:
        """檢查是否為外部連結"""
        if not file_path:
            return False
        
        file_path = file_path.strip()
        is_external = file_path.startswith(('http://', 'https://', 'ftp://', 'mailto:'))
        self.logger.debug(f"外部連結檢查: {file_path} -> {is_external}")
        return is_external
    
    def _is_html_file(self, file_path: str) -> bool:
        """檢查是否為 HTML 檔案"""
        self.logger.debug(f"檢查 HTML 檔案: {file_path}")
        if not file_path:
            return False
        
        clean_path = file_path.split('?')[0].split('#')[0]
        result = clean_path.lower().endswith(('.html', '.htm'))
        self.logger.debug(f"HTML 檔案檢查結果: {result}")
        return result
    

    
    def _parse_html_media_files(self, html_file_path: Path, folder_path: str) -> Tuple[int, int, int, str]:
        """
        解析 HTML 檔案中的媒體檔案
        
        Args:
            html_file_path: HTML 檔案路徑
            folder_path: 資料夾路徑前綴
            
        Returns:
            Tuple[int, int, int, str]: (文件數, 有效文件數, 遺失數, 遺失明細)
        """
        try:
            self.logger.info(f"開始解析 HTML 檔案: {html_file_path}")
            
            # 讀取 HTML 檔案
            with open(html_file_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            # 解析 HTML
            parser = MediaFileParser()
            parser.feed(html_content)
            media_files = parser.get_media_files()
            
            # HTML 檔案本身算 1 個文件
            html_file_count = 1
            html_valid_count = 1  # HTML 檔案存在才能解析到這裡
            
            if not media_files:
                self.logger.info(f"HTML 檔案中未找到媒體檔案: {html_file_path}")
                # 只有 HTML 檔案本身
                return html_file_count, html_valid_count, 0, ""
            
            # 檢查媒體檔案是否存在
            html_dir = html_file_path.parent
            valid_media_files = 0
            missing_files = []
            
            for media_file in media_files:
                # 計算媒體檔案的完整路徑
                media_path = html_dir / media_file
                
                # 嘗試不同的路徑組合
                possible_paths = [
                    media_path,
                    html_dir / media_file.lstrip('./'),
                    html_dir / media_file.lstrip('../').lstrip('./'),
                ]
                
                # 檢查檔案是否存在
                file_exists = any(p.exists() for p in possible_paths)
                
                if file_exists:
                    valid_media_files += 1
                    self.logger.debug(f"媒體檔案存在: {media_file}")
                else:
                    # 直接使用媒體檔案路徑
                    clean_media_file = media_file
                    missing_files.append(clean_media_file)
                    self.logger.warning(f"媒體檔案不存在: {media_file} (HTML: {html_file_path})")
            
            # 計算總計
            total_files = html_file_count + len(media_files)  # HTML 檔案 + 媒體檔案
            total_valid_files = html_valid_count + valid_media_files  # 有效的 HTML + 有效的媒體檔案
            total_missing = len(missing_files)  # 只有媒體檔案會遺失，HTML 檔案已經存在
            missing_details = "; ".join(missing_files)
            
            self.logger.info(f"HTML 分析完成: {html_file_path} - 總計:{total_files}(HTML:1+媒體:{len(media_files)}), 有效:{total_valid_files}, 遺失:{total_missing}")
            self.stats['html_files_parsed'] += 1
            
            return total_files, total_valid_files, total_missing, missing_details
            
        except UnicodeDecodeError:
            # 嘗試其他編碼
            try:
                self.logger.info(f"嘗試使用 Big5 編碼解析 HTML: {html_file_path}")
                with open(html_file_path, 'r', encoding='big5') as f:
                    html_content = f.read()
                
                parser = MediaFileParser()
                parser.feed(html_content)
                media_files = parser.get_media_files()
                
                # HTML 檔案本身算 1 個文件
                html_file_count = 1
                html_valid_count = 1
                
                if not media_files:
                    return html_file_count, html_valid_count, 0, ""
                
                html_dir = html_file_path.parent
                valid_media_files = 0
                missing_files = []
                
                for media_file in media_files:
                    media_path = html_dir / media_file
                    possible_paths = [
                        media_path,
                        html_dir / media_file.lstrip('./'),
                        html_dir / media_file.lstrip('../').lstrip('./'),
                    ]
                    
                    if any(p.exists() for p in possible_paths):
                        valid_media_files += 1
                    else:
                        # 直接使用媒體檔案路徑
                        clean_media_file = media_file
                        missing_files.append(clean_media_file)
                
                total_files = html_file_count + len(media_files)
                total_valid_files = html_valid_count + valid_media_files
                total_missing = len(missing_files)
                missing_details = "; ".join(missing_files)
                
                self.logger.info(f"HTML 分析完成 (Big5): {html_file_path} - 總計:{total_files}(HTML:1+媒體:{len(media_files)}), 有效:{total_valid_files}, 遺失:{total_missing}")
                self.stats['html_files_parsed'] += 1
                
                return total_files, total_valid_files, total_missing, missing_details
                
            except Exception as e:
                self.logger.error(f"HTML 解析失敗 (嘗試多種編碼): {html_file_path}, 錯誤: {e}")
                return 1, 0, 1, f"{html_file_path.name} #解析失敗"
        
        except Exception as e:
            self.logger.error(f"HTML 解析失敗: {html_file_path}, 錯誤: {e}")
            return 1, 0, 1, f"{html_file_path.name} #解析失敗"
    
    def _analyze_system_path(self, system_path: str, original_path: str, folder_path: str) -> Tuple[int, int, int, str]:
        """
        基於系統路徑分析檔案
        
        Args:
            system_path: 完整的系統路徑
            original_path: 原始路徑（用於遺失明細）
            folder_path: 資料夾路徑前綴
            
        Returns:
            Tuple[int, int, int, str]: (文件數, 有效文件數, 遺失數, 遺失明細)
        """
        if not system_path or not system_path.strip():
            return 0, 0, 0, ""
        
        system_path_str = system_path.strip()
        
        # 檢查是否為外部連結
        if self._is_external_link(system_path_str):
            self.logger.info(f"檢測到外部連結: {system_path_str}")
            # 外部連結不計入文件數，也不會有遺失問題
            return 0, 0, 0, ""
        
        system_path_obj = Path(system_path_str)
        
        self.logger.debug(f"分析系統路徑: {system_path_str}")
        
        # 檢查是否為 HTML 檔案
        if self._is_html_file(system_path_str):
            self.logger.info(f"檢測到 HTML 檔案: {system_path_str}")
            
            # 檢查 HTML 檔案本身是否存在
            if not system_path_obj.exists():
                self.logger.warning(f"HTML 檔案不存在: {system_path_str}")
                # 直接使用原始路徑
                return 1, 0, 1, original_path
            
            # 解析 HTML 檔案中的媒體檔案
            total_files, valid_files, missing_count, missing_details = self._parse_html_media_files(system_path_obj, folder_path)
            return total_files, valid_files, missing_count, missing_details
        
        else:
            # 非 HTML 檔案
            if system_path_obj.exists():
                self.logger.debug(f"檔案存在: {system_path_str}")
                return 1, 1, 0, ""
            else:
                self.logger.warning(f"檔案不存在: {system_path_str}")
                # 直接使用原始路徑
                return 1, 0, 1, original_path
    
    def _insert_analysis_columns(self, worksheet, positions: Dict[str, int]):
        """在路徑欄位後插入分析欄位"""
        path_info = positions.get('path', {})
        if not path_info or 'col' not in path_info:
            self.logger.error("未找到路徑欄位，無法插入分析欄位")
            return {}
        
        path_col = path_info['col']
        path_row = path_info['row']
        
        # 插入新欄位
        insert_col = path_col + 1
        
        # 插入5個新欄位
        for i in range(5):
            worksheet.insert_cols(insert_col)
        
        # 設定新欄位標題
        new_columns = {
            'system_path': insert_col,
            'file_count': insert_col + 1,
            'valid_count': insert_col + 2,
            'missing_count': insert_col + 3,
            'missing_details': insert_col + 4
        }
        
        # 設定標題（在與"路徑"同一行）
        worksheet.cell(row=path_row, column=new_columns['system_path']).value = "系統路徑"
        worksheet.cell(row=path_row, column=new_columns['file_count']).value = "文件數"
        worksheet.cell(row=path_row, column=new_columns['valid_count']).value = "有效文件數"
        worksheet.cell(row=path_row, column=new_columns['missing_count']).value = "遺失數"
        worksheet.cell(row=path_row, column=new_columns['missing_details']).value = "遺失明細"
        
        # 設定標題樣式
        header_font = Font(bold=True)
        for col in new_columns.values():
            cell = worksheet.cell(row=path_row, column=col)
            cell.font = header_font
        
        # 調整欄寬
        worksheet.column_dimensions[get_column_letter(new_columns['system_path'])].width = 50
        worksheet.column_dimensions[get_column_letter(new_columns['file_count'])].width = 12
        worksheet.column_dimensions[get_column_letter(new_columns['valid_count'])].width = 15
        worksheet.column_dimensions[get_column_letter(new_columns['missing_count'])].width = 12
        worksheet.column_dimensions[get_column_letter(new_columns['missing_details'])].width = 60
        
        self.logger.info(f"插入分析欄位完成: {new_columns}")
        return new_columns
    
    def _analyze_worksheet(self, worksheet, sheet_name: str) -> bool:
        """分析單個工作表"""
        try:
            self.logger.info(f"開始分析工作表: {sheet_name}")
            
            # 獲取工作表資訊（從固定的A2和A3讀取）
            sheet_info = self._get_sheet_info(worksheet)
            resource_path = sheet_info['resource_path']
            folder_path = sheet_info['folder_path']
            
            if not folder_path:
                self.logger.warning(f"工作表 {sheet_name} 未找到資料夾路徑")
                return False
            
            self.logger.info(f"工作表 {sheet_name} - 資源庫路徑: {resource_path}, 資料夾路徑: {folder_path}")
            
            # 找到欄位位置
            positions = self._find_column_positions(worksheet)
            if not positions.get('path'):
                self.logger.warning(f"工作表 {sheet_name} 未找到路徑欄位")
                return False
            
            # 找到資料開始行號
            data_start_row = self._find_data_start_row(worksheet, positions)
            
            # 插入分析欄位
            new_columns = self._insert_analysis_columns(worksheet, positions)
            if not new_columns:
                return False
            
            # 分析每一行的路徑
            path_col = positions['path']['col']
            paths_found = 0
            paths_modified = 0
            
            for row_num in range(data_start_row, worksheet.max_row + 1):
                original_path = worksheet.cell(row=row_num, column=path_col).value
                
                if original_path and str(original_path).strip():
                    original_path_str = str(original_path).strip()
                    paths_found += 1
                    
                    self.logger.debug(f"處理第 {row_num} 行路徑: {original_path_str}")
                    
                    # 修改路徑欄位的值：資源庫路徑&原路徑
                    if resource_path and not self._is_external_link(original_path_str):
                        new_path = f"{resource_path}{original_path_str}"
                        worksheet.cell(row=row_num, column=path_col).value = new_path
                        paths_modified += 1
                        self.logger.debug(f"修改路徑: {original_path_str} -> {new_path}")
                    
                    # 建構系統路徑（維持原邏輯：資料夾路徑/原路徑）
                    if self._is_external_link(original_path_str):
                        system_path = original_path_str
                    else:
                        system_path = str(Path(folder_path) / original_path_str)
                    
                    # 分析系統路徑
                    file_count, valid_count, missing_count, missing_details = self._analyze_system_path(
                        system_path, original_path_str, folder_path
                    )
                    
                    # 寫入分析結果
                    worksheet.cell(row=row_num, column=new_columns['system_path']).value = system_path
                    worksheet.cell(row=row_num, column=new_columns['file_count']).value = file_count
                    worksheet.cell(row=row_num, column=new_columns['valid_count']).value = valid_count
                    worksheet.cell(row=row_num, column=new_columns['missing_count']).value = missing_count
                    worksheet.cell(row=row_num, column=new_columns['missing_details']).value = missing_details
                    
                    self.stats['paths_analyzed'] += 1
                    self.logger.debug(f"分析完成 - 行{row_num}: {original_path_str} -> 文件數:{file_count}, 有效:{valid_count}, 遺失:{missing_count}")
            
            self.logger.info(f"工作表 {sheet_name} 分析完成，共處理 {paths_found} 個路徑，修改 {paths_modified} 個路徑")
            self.stats['sheets_processed'] += 1
            self.stats['paths_modified'] += paths_modified
            return True
            
        except Exception as e:
            self.logger.error(f"分析工作表 {sheet_name} 失敗: {e}")
            import traceback
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")
            self.stats['errors'] += 1
            return False
    
    def analyze_excel(self) -> bool:
        """分析 Excel 檔案"""
        try:
            # 選擇 Excel 檔案
            if not self._select_excel_file():
                return False
            
            # 生成時間戳和輸出檔名
            self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_filename = f"course_structures_analyzed_{self.timestamp}.xlsx"
            self.output_file = self.output_dir / output_filename
            
            self.logger.info(f"開始分析 Excel 檔案: {self.selected_excel_file.name}")
            
            # 載入原始 Excel 檔案
            workbook = openpyxl.load_workbook(self.selected_excel_file)
            
            # 分析每個工作表
            for sheet_name in workbook.sheetnames:
                print(f"正在分析工作表: {sheet_name}")
                self._analyze_worksheet(workbook[sheet_name], sheet_name)
            
            # 儲存結果
            workbook.save(self.output_file)
            workbook.close()
            
            self.logger.info(f"分析完成，結果儲存至: {self.output_file}")
            return True
            
        except Exception as e:
            self.logger.error(f"分析 Excel 檔案失敗: {e}")
            import traceback
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")
            return False
    
    def print_summary(self):
        """輸出處理摘要"""
        print("\n" + "="*50)
        print("📊 Excel 路徑分析作業完成")
        print("="*50)
        print(f"來源 Excel 檔案: {self.selected_excel_file.name if self.selected_excel_file else 'N/A'}")
        print(f"處理工作表數: {self.stats['sheets_processed']}")
        print(f"分析路徑數: {self.stats['paths_analyzed']}")
        print(f"修改路徑數: {self.stats['paths_modified']}")
        print(f"解析 HTML 檔案數: {self.stats['html_files_parsed']}")
        if self.stats['errors'] > 0:
            print(f"錯誤數: {self.stats['errors']}")
        
        if self.output_file:
            print(f"\n📁 分析結果檔案: {self.output_file.name}")
            print(f"📂 檔案位置: {self.output_file.parent.absolute()}")


def main():
    """主函數"""
    print("🔍 Excel 路徑分析器 - 調整版")
    print("="*30)
    
    # 檢查依賴
    try:
        import openpyxl
    except ImportError:
        print("❌ 缺少必要的依賴套件：openpyxl")
        print("請執行：pip install openpyxl")
        return False
    
    # 建立分析器並執行
    analyzer = ExcelPathAnalyzer()
    success = analyzer.analyze_excel()
    
    # 輸出摘要
    analyzer.print_summary()
    
    if success:
        print(f"\n🎉 路徑分析完成！")
        print(f"📁 分析結果檔案：{analyzer.output_file.name}")
        print(f"📂 檔案位置：{analyzer.output_dir.absolute()}")
        print(f"\n✨ 主要調整：")
        print(f"• 路徑欄位已更新為：資源庫路徑/原路徑")
        print(f"• 遺失明細移除資料夾路徑前綴")
        print(f"• 系統路徑維持原邏輯：資料夾路徑/原路徑")
    else:
        print(f"\n❌ 路徑分析過程中發生錯誤，請檢查日誌檔案")
    
    return success


if __name__ == "__main__":
    main()