#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 課程結構統計報告生成器 - 優化版本
分析已生成的 Excel 課程結構檔案，產生統計報告
修正章節和單元的動態欄位識別邏輯
"""

import os
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class ExcelStatistics:
    """Excel 統計報告生成器類別"""
    
    def __init__(self, output_dir: str = "to_be_executed"):
        """
        初始化統計報告生成器
        
        Args:
            output_dir: 輸出目錄路徑
        """
        self.output_dir = Path(output_dir)
        self.report_file = None
        self.timestamp = None
        self.selected_excel_file = None
        
        # 統計資料
        self.course_stats = {}  # 每個課程的統計
        self.overall_stats = {   # 總體統計
            'total_courses': 0,
            'total_chapters': 0,
            'total_units': 0,
            'unit_breakdown': {},
            'total_activities': 0,
            'activity_types': {},
            'total_valid_files': 0,      # 新增：總有效文件數
            'total_missing_files': 0,    # 新增：總遺失文件數
            'missing_details': []        # 新增：遺失明細
        }
        
        self.stats = {
            'sheets_analyzed': 0,
            'errors': 0
        }
        
        # 設定日誌
        self._setup_logging()
    
    def _setup_logging(self):
        """設定日誌系統"""
        # 讓 log 儲存在 ../log 資料夾中
        log_filename = f"excel_statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        log_dir = Path("../log")
        log_dir.mkdir(parents=True, exist_ok=True)
        log_path = log_dir / log_filename

        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_path, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )

        self.logger = logging.getLogger(__name__)
        self.logger.info("=== 開始 Excel 統計報告生成作業 ===")
        self.logger.info(f"輸出目錄: {self.output_dir.absolute()}")
        self.logger.info(f"Log 檔案儲存於: {log_path}")
    
    def _create_output_directory(self):
        """建立輸出目錄"""
        try:
            self.output_dir.mkdir(parents=True, exist_ok=True)
            self.logger.info(f"輸出目錄已準備: {self.output_dir}")
        except Exception as e:
            self.logger.error(f"無法建立輸出目錄: {e}")
            raise
    
    def _find_excel_files(self) -> List[Path]:
        """
        找到所有 course_structures_analyzed 開頭的 Excel 檔案
        
        Returns:
            List[Path]: Excel 檔案路徑列表
        """
        if not self.output_dir.exists():
            return []
        
        # 找到所有 course_structures_analyzed 開頭的 Excel 檔案
        pattern = "course_structures_analyzed_*.xlsx"
        excel_files = []
        
        for file_path in self.output_dir.glob(pattern):
            # 排除包含 report 的檔案
            if "report" not in file_path.name.lower():
                excel_files.append(file_path)
        
        return sorted(excel_files, key=lambda x: x.stat().st_mtime, reverse=True)
    
    def _select_excel_file(self) -> bool:
        """
        讓用戶選擇要分析的 Excel 檔案
        
        Returns:
            bool: 是否成功選擇檔案
        """
        excel_files = self._find_excel_files()
        
        if not excel_files:
            print("❌ 在 to_be_executed 目錄中未找到 course_structures_analyzed 開頭的 Excel 檔案")
            return False
        
        print(f"\n找到 {len(excel_files)} 個可分析的 Excel 檔案：")
        print("=" * 50)
        
        for i, file_path in enumerate(excel_files, 1):
            file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
            file_size = file_path.stat().st_size
            print(f"{i}. {file_path.name}")
            print(f"   修改時間: {file_time.strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"   檔案大小: {file_size:,} bytes")
            print()
        
        # 用戶選擇
        while True:
            try:
                choice = input(f"請選擇要分析的 Excel 檔案 (1-{len(excel_files)}) [預設: 1]: ").strip()
                if not choice:
                    choice = "1"
                
                index = int(choice) - 1
                if 0 <= index < len(excel_files):
                    self.selected_excel_file = excel_files[index]
                    print(f"✅ 已選擇: {self.selected_excel_file.name}")
                    return True
                else:
                    print(f"❌ 請輸入 1 到 {len(excel_files)} 之間的數字")
            except ValueError:
                print("❌ 請輸入有效的數字")
            except KeyboardInterrupt:
                print("\n❌ 操作已取消")
                return False
    
    def _analyze_column_headers(self, worksheet) -> Dict[str, int]:
        """
        分析工作表的欄位標題，找出動態欄位位置
        
        Args:
            worksheet: 工作表物件
            
        Returns:
            Dict[str, int]: 欄位名稱到欄位編號的映射
        """
        col_mapping = {}
        
        # 檢查第6行（層級標題行）
        level_headers_row = 6
        max_col = worksheet.max_column
        
        # 讀取層級標題
        for col in range(1, max_col + 1):
            cell_value = worksheet.cell(row=level_headers_row, column=col).value
            if cell_value and str(cell_value).strip():
                header_name = str(cell_value).strip()
                col_mapping[header_name] = col
        
        # 檢查第5行（系統欄位標題行）
        system_headers_row = 5
        for col in range(1, max_col + 1):
            cell_value = worksheet.cell(row=system_headers_row, column=col).value
            if cell_value and str(cell_value).strip():
                header_name = str(cell_value).strip()
                # 只添加不在層級標題中的欄位
                if header_name not in col_mapping:
                    col_mapping[header_name] = col
        
        return col_mapping
    
    def _extract_filenames_from_missing_details(self, missing_details: str) -> List[str]:
        """
        從遺失明細中提取檔案名稱
        
        Args:
            missing_details: 遺失明細文字
            
        Returns:
            List[str]: 檔案名稱列表
        """
        import os
        
        filenames = []
        
        # 分割多個檔案（如果有的話）
        # 遺失明細可能包含多個檔案，用分號、逗號或換行分隔
        detail_parts = missing_details.replace(';', '\n').replace(',', '\n').split('\n')
        
        for part in detail_parts:
            part = part.strip()
            if not part:
                continue
                
            # 移除狀態說明（如 "# 檔案不存在"）
            if '#' in part:
                part = part.split('#')[0].strip()
            
            # 如果包含路徑，提取檔案名稱
            if '/' in part or '\\' in part:
                # 使用 os.path.basename 提取檔案名稱
                filename = os.path.basename(part.strip())
            else:
                # 直接是檔案名稱
                filename = part.strip()
            
            if filename:
                filenames.append(filename)
        
        return filenames
    
    def _analyze_excel_sheet(self, worksheet, sheet_name: str) -> bool:
        """
        分析單個 Excel 工作表 - 讀取 analyzed 檔案中已分析的資料
        
        Args:
            worksheet: 工作表物件
            sheet_name: 工作表名稱
            
        Returns:
            bool: 分析是否成功
        """
        try:
            # 初始化課程統計
            course_stat = {
                'chapters': 0,
                'units': 0,
                'unit_breakdown': {},
                'activities': 0,
                'activity_types': {},
                'valid_files': 0,        # 新增：有效文件數
                'missing_files': 0,      # 新增：遺失文件數
                'missing_details': []    # 新增：遺失明細
            }
            
            # 分析欄位標題
            col_mapping = self._analyze_column_headers(worksheet)
            self.logger.info(f"工作表 {sheet_name} 欄位映射: {col_mapping}")
            
            # 識別關鍵欄位
            chapter_col = col_mapping.get('章節', None)
            unit_cols = []  # 儲存所有單元欄位
            activity_col = col_mapping.get('學習活動', None)
            type_col = col_mapping.get('類型', None)
            valid_files_col = col_mapping.get('有效文件數', None)
            missing_files_col = col_mapping.get('遺失數', None)
            missing_details_col = col_mapping.get('遺失明細', None)
            
            # 尋找所有單元欄位
            for header_name, col_num in col_mapping.items():
                if header_name.startswith('單元') and header_name[2:].isdigit():
                    unit_cols.append((header_name, col_num))
            
            # 按單元編號排序
            unit_cols.sort(key=lambda x: int(x[0][2:]))
            
            self.logger.info(f"工作表 {sheet_name} 關鍵欄位: 章節={chapter_col}, 單元={unit_cols}, 活動={activity_col}")
            
            # 確認必要欄位存在
            if type_col is None:
                self.logger.warning(f"工作表 {sheet_name} 缺少必要欄位: 類型")
                return False
            
            # 資料開始行
            data_start_row = 7
            
            # 用於追蹤已統計的項目（避免重複計算）
            counted_chapters = set()
            counted_units = {unit_name: set() for unit_name, _ in unit_cols}
            counted_missing_files = set()  # 用於追蹤已統計的遺失檔案名稱
            
            # 分析每一行資料
            for row_num in range(data_start_row, worksheet.max_row + 1):
                # 檢查是否為空行
                has_data = False
                for col in range(1, worksheet.max_column + 1):
                    if worksheet.cell(row=row_num, column=col).value:
                        has_data = True
                        break
                
                if not has_data:
                    continue
                
                # 統計章節
                if chapter_col:
                    chapter_value = worksheet.cell(row=row_num, column=chapter_col).value
                    if chapter_value and str(chapter_value).strip():
                        chapter_name = str(chapter_value).strip()
                        if chapter_name not in counted_chapters:
                            counted_chapters.add(chapter_name)
                            course_stat['chapters'] += 1
                            self.overall_stats['total_chapters'] += 1
                
                # 統計單元
                for unit_name, unit_col in unit_cols:
                    unit_value = worksheet.cell(row=row_num, column=unit_col).value
                    if unit_value and str(unit_value).strip():
                        unit_content = str(unit_value).strip()
                        unit_key = f"{chapter_name if 'chapter_name' in locals() else 'unknown'}-{unit_content}"
                        
                        if unit_key not in counted_units[unit_name]:
                            counted_units[unit_name].add(unit_key)
                            course_stat['units'] += 1
                            course_stat['unit_breakdown'][unit_name] = course_stat['unit_breakdown'].get(unit_name, 0) + 1
                            self.overall_stats['total_units'] += 1
                            self.overall_stats['unit_breakdown'][unit_name] = self.overall_stats['unit_breakdown'].get(unit_name, 0) + 1
                
                # 讀取活動資料
                type_value = None
                valid_files = 0
                missing_files = 0
                missing_details = ""
                
                # 讀取類型
                if type_col:
                    type_cell = worksheet.cell(row=row_num, column=type_col)
                    type_value = type_cell.value if type_cell.value else None
                
                # 讀取有效文件數
                if valid_files_col:
                    valid_cell = worksheet.cell(row=row_num, column=valid_files_col)
                    if valid_cell.value is not None:
                        try:
                            valid_files = int(valid_cell.value) if valid_cell.value != '' else 0
                        except (ValueError, TypeError):
                            valid_files = 0
                
                # 讀取遺失文件數
                if missing_files_col:
                    missing_cell = worksheet.cell(row=row_num, column=missing_files_col)
                    if missing_cell.value is not None:
                        try:
                            missing_files = int(missing_cell.value) if missing_cell.value != '' else 0
                        except (ValueError, TypeError):
                            missing_files = 0
                
                # 讀取遺失明細
                if missing_details_col:
                    details_cell = worksheet.cell(row=row_num, column=missing_details_col)
                    missing_details = str(details_cell.value) if details_cell.value else ""
                
                # 如果有活動類型，統計學習活動
                if type_value and str(type_value).strip():
                    activity_type = str(type_value).strip()
                    course_stat['activities'] += 1
                    course_stat['activity_types'][activity_type] = course_stat['activity_types'].get(activity_type, 0) + 1
                    self.overall_stats['total_activities'] += 1
                    self.overall_stats['activity_types'][activity_type] = self.overall_stats['activity_types'].get(activity_type, 0) + 1
                    
                    # 統計文件狀態
                    course_stat['valid_files'] += valid_files
                    self.overall_stats['total_valid_files'] += valid_files
                    
                    # 處理遺失文件統計（按檔案名稱去重）
                    if missing_files > 0 and missing_details.strip():
                        # 從遺失明細中提取檔案名稱
                        missing_detail_text = missing_details.strip()
                        # 遺失明細格式通常是：檔案名稱 # 檔案不存在 或 路徑/檔案名稱
                        
                        # 分析遺失明細中的檔案名稱
                        extracted_filenames = self._extract_filenames_from_missing_details(missing_detail_text)
                        
                        # 統計不重複的檔案名稱
                        unique_missing_count = 0
                        for filename in extracted_filenames:
                            if filename not in counted_missing_files:
                                counted_missing_files.add(filename)
                                unique_missing_count += 1
                        
                        # 使用去重後的數量
                        course_stat['missing_files'] += unique_missing_count
                        self.overall_stats['total_missing_files'] += unique_missing_count
                        
                        # 加入遺失明細（保留原始格式）
                        if unique_missing_count > 0:
                            # 獲取活動名稱
                            activity_name = "未知活動"
                            if activity_col:
                                activity_cell = worksheet.cell(row=row_num, column=activity_col)
                                if activity_cell.value and str(activity_cell.value).strip():
                                    activity_name = str(activity_cell.value).strip()
                            
                            missing_detail = {
                                'course': sheet_name,
                                'activity': activity_name,
                                'type': activity_type,
                                'path': missing_details.strip()
                            }
                            
                            course_stat['missing_details'].append(missing_detail)
                            self.overall_stats['missing_details'].append(missing_detail)
                    else:
                        # 沒有遺失明細但有遺失數量的情況
                        course_stat['missing_files'] += missing_files
                        self.overall_stats['total_missing_files'] += missing_files
            
            self.course_stats[sheet_name] = course_stat
            self.overall_stats['total_courses'] += 1
            
            # 日誌輸出統計結果
            total_missing_in_details = len(course_stat['missing_details'])
            self.logger.info(f"成功分析工作表: {sheet_name}")
            self.logger.info(f"  - 章節: {course_stat['chapters']}個")
            self.logger.info(f"  - 單元: {course_stat['units']}個")
            self.logger.info(f"  - 學習活動: {course_stat['activities']}個")
            self.logger.info(f"  - 有效文件: {course_stat['valid_files']}個")
            self.logger.info(f"  - 遺失文件: {course_stat['missing_files']}個 (去重後)")
            self.logger.info(f"  - 遺失明細項目: {total_missing_in_details}項")
            
            return True
            
        except Exception as e:
            self.stats['errors'] += 1
            self.logger.error(f"分析工作表失敗 {sheet_name}: {e}")
            import traceback
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")
            return False
    
    def _analyze_excel_file(self) -> bool:
        """
        分析選擇的 Excel 檔案
        
        Returns:
            bool: 分析是否成功
        """
        try:
            self.logger.info(f"開始分析 Excel 檔案: {self.selected_excel_file.name}")
            
            # 載入 Excel 檔案
            workbook = openpyxl.load_workbook(self.selected_excel_file)
            
            # 分析每個工作表
            for sheet_name in workbook.sheetnames:
                print(f"正在分析工作表: {sheet_name}")
                worksheet = workbook[sheet_name]
                
                success = self._analyze_excel_sheet(worksheet, sheet_name)
                if success:
                    self.stats['sheets_analyzed'] += 1
            
            workbook.close()
            
            self.logger.info(f"Excel 檔案分析完成，共分析 {self.stats['sheets_analyzed']} 個工作表")
            return True
            
        except Exception as e:
            self.logger.error(f"分析 Excel 檔案失敗: {e}")
            return False
    
    def _create_report_excel(self) -> bool:
        """
        建立統計分析報告 Excel
        
        Returns:
            bool: 建立是否成功
        """
        try:
            self.logger.info("開始建立統計報告...")
            
            # 建立報告工作簿
            report_workbook = openpyxl.Workbook()
            self.logger.info("報告工作簿已建立")
            
            # 移除預設工作表
            default_sheet = report_workbook.active
            report_workbook.remove(default_sheet)
            self.logger.info("預設工作表已移除")
            
            # 建立匯總統計工作表
            summary_sheet = report_workbook.create_sheet(title="匯總統計")
            self.logger.info("匯總統計工作表已建立")
            
            self._create_summary_sheet(summary_sheet)
            self.logger.info("匯總統計工作表內容已建立")
            
            # 建立各課程統計工作表
            detail_sheet = report_workbook.create_sheet(title="各課程統計")
            self.logger.info("各課程統計工作表已建立")
            
            self._create_detail_sheet(detail_sheet)
            self.logger.info("各課程統計工作表內容已建立")
            
            # 建立遺失明細工作表
            missing_sheet = report_workbook.create_sheet(title="遺失明細")
            self.logger.info("遺失明細工作表已建立")
            
            self._create_missing_sheet(missing_sheet)
            self.logger.info("遺失明細工作表內容已建立")
            
            # 儲存報告檔案
            self.logger.info(f"準備儲存報告檔案: {self.report_file}")
            report_workbook.save(self.report_file)
            
            self.logger.info(f"統計報告已儲存: {self.report_file}")
            return True
            
        except Exception as e:
            import traceback
            self.logger.error(f"建立統計報告失敗: {e}")
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")
            return False
    
    def _create_summary_sheet(self, worksheet):
        """建立匯總統計工作表"""
        try:
            self.logger.info("開始建立匯總統計工作表內容...")
            
            # 標題
            worksheet['A1'] = "課程結構分析報告 - 匯總統計"
            worksheet['A1'].font = Font(bold=True, size=16)
            self.logger.info("標題已設定")
            
            worksheet['A2'] = f"生成時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            worksheet['A3'] = f"來源檔案：{self.selected_excel_file.name}"
            worksheet['A4'] = f"統計報告檔案：{self.report_file.name}"
            self.logger.info("基本資訊已設定")
            
            # 匯總統計
            row = 6
            worksheet[f'A{row}'] = "總體統計"
            worksheet[f'A{row}'].font = Font(bold=True, size=14)
            self.logger.info("總體統計標題已設定")
            
            row += 1
            stats_data = [
                ("總課程數", self.overall_stats['total_courses']),
                ("總章節數", self.overall_stats['total_chapters']),
                ("總單元數", self.overall_stats['total_units']),
                ("總學習活動數", self.overall_stats['total_activities']),
                ("總有效文件數", self.overall_stats['total_valid_files']),
                ("總遺失文件數", self.overall_stats['total_missing_files'])
            ]
            
            for label, value in stats_data:
                worksheet[f'A{row}'] = label
                worksheet[f'B{row}'] = f"{value}個"
                worksheet[f'A{row}'].font = Font(bold=True)
                # 遺失文件數用紅色標記
                if "遺失" in label and value > 0:
                    worksheet[f'B{row}'].font = Font(color="FF0000", bold=True)
                row += 1
            self.logger.info("基本統計數據已設定")
            
            # 文件狀態統計
            if self.overall_stats['total_activities'] > 0:
                row += 1
                valid_rate = (self.overall_stats['total_valid_files'] / self.overall_stats['total_activities']) * 100
                missing_rate = (self.overall_stats['total_missing_files'] / self.overall_stats['total_activities']) * 100
                
                worksheet[f'A{row}'] = "文件狀態統計"
                worksheet[f'A{row}'].font = Font(bold=True)
                row += 1
                worksheet[f'B{row}'] = f"有效率: {valid_rate:.1f}%"
                row += 1
                worksheet[f'B{row}'] = f"遺失率: {missing_rate:.1f}%"
                if missing_rate > 0:
                    worksheet[f'B{row}'].font = Font(color="FF0000")
                row += 1
                self.logger.info("文件狀態統計已設定")
            
            # 單元細分
            if self.overall_stats['unit_breakdown']:
                row += 1
                worksheet[f'A{row}'] = "單元細分"
                worksheet[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for unit_name, count in sorted(self.overall_stats['unit_breakdown'].items()):
                    worksheet[f'B{row}'] = f"- {unit_name}"
                    worksheet[f'C{row}'] = f"{count}個"
                    row += 1
                self.logger.info("單元細分已設定")
            
            # 學習活動統計
            row += 1
            worksheet[f'A{row}'] = "活動類型細分"
            worksheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for activity_type, count in sorted(self.overall_stats['activity_types'].items(), key=lambda x: x[1], reverse=True):
                worksheet[f'B{row}'] = f"- {activity_type}"
                worksheet[f'C{row}'] = f"{count}個"
                row += 1
            self.logger.info("活動類型統計已設定")
            
            # 調整欄寬
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 15
            self.logger.info("匯總統計工作表完成")
            
        except Exception as e:
            import traceback
            self.logger.error(f"建立匯總統計工作表失敗: {e}")
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")
            raise
    
    def _create_detail_sheet(self, worksheet):
        """建立各課程統計工作表"""
        try:
            self.logger.info("開始建立各課程統計工作表內容...")
            
            # 標題
            worksheet['A1'] = "各課程詳細統計"
            worksheet['A1'].font = Font(bold=True, size=16)
            self.logger.info("標題已設定")
            
            # 表頭
            headers = ['課程名稱', '章節數', '單元數', '學習活動數', '有效文件數', '遺失文件數', '遺失率', '活動類型細分']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
            self.logger.info("表頭已設定")
            
            # 資料
            row = 4
            for course_name, stats in self.course_stats.items():
                worksheet[f'A{row}'] = course_name
                worksheet[f'B{row}'] = f"{stats['chapters']}個"
                worksheet[f'C{row}'] = f"{stats['units']}個"
                worksheet[f'D{row}'] = f"{stats['activities']}個"
                worksheet[f'E{row}'] = f"{stats['valid_files']}個"
                worksheet[f'F{row}'] = f"{stats['missing_files']}個"
                
                # 遺失率計算
                if stats['activities'] > 0:
                    missing_rate = (stats['missing_files'] / stats['activities']) * 100
                    worksheet[f'G{row}'] = f"{missing_rate:.1f}%"
                    if missing_rate > 0:
                        worksheet[f'F{row}'].font = Font(color="FF0000")
                        worksheet[f'G{row}'].font = Font(color="FF0000")
                else:
                    worksheet[f'G{row}'] = "0%"
                
                # 活動類型細分
                activity_breakdown = ", ".join([f"{k}:{v}個" for k, v in sorted(stats['activity_types'].items(), key=lambda x: x[1], reverse=True)])
                worksheet[f'H{row}'] = activity_breakdown if activity_breakdown else "-"
                
                row += 1
            
            self.logger.info("資料已設定")
            
            # 調整欄寬
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 12
            worksheet.column_dimensions['C'].width = 12
            worksheet.column_dimensions['D'].width = 15
            worksheet.column_dimensions['E'].width = 15
            worksheet.column_dimensions['F'].width = 15
            worksheet.column_dimensions['G'].width = 12
            worksheet.column_dimensions['H'].width = 40
            self.logger.info("各課程統計工作表完成")
            
        except Exception as e:
            import traceback
            self.logger.error(f"建立各課程統計工作表失敗: {e}")
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")
            raise
    
    def _create_missing_sheet(self, worksheet):
        """建立遺失明細工作表"""
        try:
            self.logger.info("開始建立遺失明細工作表內容...")
            
            # 標題
            worksheet['A1'] = "遺失文件明細"
            worksheet['A1'].font = Font(bold=True, size=16)
            self.logger.info("標題已設定")
            
            # 表頭
            headers = ['課程名稱', '學習活動', '活動類型', '路徑狀態']
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
            self.logger.info("表頭已設定")
            
            # 資料
            if self.overall_stats['missing_details']:
                row = 4
                for detail in self.overall_stats['missing_details']:
                    worksheet[f'A{row}'] = detail['course']
                    worksheet[f'B{row}'] = detail['activity']
                    worksheet[f'C{row}'] = detail['type']
                    worksheet[f'D{row}'] = detail['path']
                    
                    # 遺失項目用紅色標記
                    for col in range(1, 5):
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = Font(color="FF0000")
                    
                    row += 1
                
                self.logger.info(f"遺失明細已設定，共 {len(self.overall_stats['missing_details'])} 項")
            else:
                worksheet['A4'] = "🎉 沒有遺失的文件！"
                worksheet['A4'].font = Font(color="00AA00", bold=True)
                self.logger.info("沒有遺失文件")
            
            # 調整欄寬
            worksheet.column_dimensions['A'].width = 35
            worksheet.column_dimensions['B'].width = 40
            worksheet.column_dimensions['C'].width = 20
            worksheet.column_dimensions['D'].width = 50
            self.logger.info("遺失明細工作表完成")
            
        except Exception as e:
            import traceback
            self.logger.error(f"建立遺失明細工作表失敗: {e}")
            self.logger.error(f"詳細錯誤: {traceback.format_exc()}")
            raise
    
    def generate_statistics(self) -> bool:
        """
        生成統計報告
        
        Returns:
            bool: 生成是否成功
        """
        try:
            # 建立輸出目錄
            self._create_output_directory()
            
            # 選擇要分析的 Excel 檔案
            if not self._select_excel_file():
                return False
            
            # 分析 Excel 檔案
            if not self._analyze_excel_file():
                return False
            
            # 生成統一的時間戳
            self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # 生成報告檔名
            report_filename = f"course_structures_report_{self.timestamp}.xlsx"
            self.report_file = self.output_dir / report_filename
            
            # 建立統計分析報告
            report_success = self._create_report_excel()
            
            if report_success:
                self.logger.info(f"統計報告已儲存: {self.report_file}")
            
            return report_success
            
        except Exception as e:
            self.logger.error(f"生成統計報告時發生錯誤: {e}")
            return False
    
    def print_summary(self):
        """輸出處理摘要"""
        print("\n" + "="*50)
        print("📊 Excel 統計報告生成作業完成")
        print("="*50)
        print(f"來源 Excel 檔案: {self.selected_excel_file.name if self.selected_excel_file else 'N/A'}")
        print(f"分析工作表數: {self.stats['sheets_analyzed']}")
        print(f"總課程數: {self.overall_stats['total_courses']}")
        print(f"總章節數: {self.overall_stats['total_chapters']}")
        print(f"總單元數: {self.overall_stats['total_units']}")
        print(f"總學習活動數: {self.overall_stats['total_activities']}")
        print(f"總有效文件數: {self.overall_stats['total_valid_files']}")
        print(f"總遺失文件數: {self.overall_stats['total_missing_files']}")
        
        if self.overall_stats['total_activities'] > 0:
            valid_rate = (self.overall_stats['total_valid_files'] / self.overall_stats['total_activities']) * 100
            missing_rate = (self.overall_stats['total_missing_files'] / self.overall_stats['total_activities']) * 100
            print(f"有效率: {valid_rate:.1f}%")
            print(f"遺失率: {missing_rate:.1f}%")
        
        if self.stats['errors'] > 0:
            print(f"錯誤數: {self.stats['errors']}")
        
        if self.report_file:
            print(f"\n📊 統計分析報告: {self.report_file.name}")
            print(f"📂 檔案位置: {self.report_file.parent.absolute()}")


def main():
    """主函數"""
    print("📊 Excel 課程結構統計報告生成器 - 優化版本")
    print("✨ 特色：單一課程內相同檔案名稱的遺失文件只計算一次")
    print("="*55)
    
    # 檢查依賴
    try:
        import openpyxl
    except ImportError:
        print("❌ 缺少必要的依賴套件：openpyxl")
        print("請執行：pip install openpyxl")
        return False
    
    # 建立統計報告生成器並執行
    statistics_generator = ExcelStatistics()
    success = statistics_generator.generate_statistics()
    
    # 輸出摘要
    statistics_generator.print_summary()
    
    if success:
        print(f"\n🎉 統計報告生成完成！")
        print(f"📊 統計分析報告：{statistics_generator.report_file.name}")
        print(f"📂 檔案位置：{statistics_generator.output_dir.absolute()}")
        
        # 顯示重要統計資訊
        if statistics_generator.overall_stats['total_missing_files'] > 0:
            print(f"\n⚠️  注意：發現 {statistics_generator.overall_stats['total_missing_files']} 個遺失文件")
            print("請查看報告中的「遺失明細」工作表了解詳情")
        else:
            print(f"\n✅ 所有文件都有效，沒有遺失！")
    else:
        print(f"\n❌ 生成統計報告過程中發生錯誤，請檢查日誌檔案")
    
    return success


if __name__ == "__main__":
    main()