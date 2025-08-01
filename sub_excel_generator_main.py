#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 課程結構生成器 - 主要生成器
掃描 4_manifest_structures 資料夾，將 JSON 課程結構轉換為 Excel 檔案
"""

import os
import json
import logging
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class ExcelGenerator:
    """Excel 生成器類別"""
    
    def __init__(self, json_dir: str = "4_manifest_structures", output_dir: str = "5_to_be_executed"):
        """
        初始化生成器
        
        Args:
            json_dir: JSON 檔案目錄路徑
            output_dir: 輸出目錄路徑
        """
        self.json_dir = Path(json_dir)
        self.output_dir = Path(output_dir)
        self.output_file = None
        self.timestamp = None  # 統一的時間戳
        
        # 路徑查找資料
        self.path_lookup = {}  # 儲存從 Excel 讀取的路徑資訊
        
        # 支援的影音檔案格式
        self.video_extensions = {
            # 影片格式
            'mpg', 'mpeg', 'mp4', 'mkv', 'avi', 'mov', 'wmv', 'flv', 'webm', 'm3u8'
        }
        
        self.audio_extensions = {
            # 音訊格式
            'mp3', 'wav', 'aac', 'ogg', 'flac', 'wma', 'midi', 'mid'
        }
        
        # 支援的圖片格式
        self.image_extensions = {
            'jpg', 'jpeg', 'png', 'gif', 'bmp', 'svg', 'webp', 'tiff', 'ico'
        }
        
        # 支援的PDF格式
        self.pdf_extensions = {'pdf'}
        
        self.stats = {
            'json_files_processed': 0,
            'sheets_created': 0,
            'items_processed': 0,
            'errors': 0,
            'fuzzy_matches': 0  # 新增：模糊查找成功次數
        }
        
        # 設定日誌
        self._setup_logging()
        
        # 載入路徑查找資料
        self._load_path_lookup()
    
    def _setup_logging(self):
        """設定日誌系統"""
        # 讓 log 儲存在 ../log 資料夾中
        log_filename = f"excel_generation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        log_dir = Path("../log")
        log_dir.mkdir(parents=True, exist_ok=True)
        log_path = log_dir / log_filename

        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(log_path, encoding='utf-8'),
                logging.StreamHandler()
            ]
        )

        self.logger = logging.getLogger(__name__)
        self.logger.info("=== 開始 Excel 生成作業 ===")
        self.logger.info(f"JSON 目錄: {self.json_dir.absolute()}")
        self.logger.info(f"輸出目錄: {self.output_dir.absolute()}")
        self.logger.info(f"Log 檔案儲存於: {log_path}")
    
    def _load_path_lookup(self):
        """載入路徑查找資料從 4_資源庫路徑_補充.xlsx"""
        lookup_file = Path("4_資源庫路徑_補充.xlsx")
        
        if not lookup_file.exists():
            self.logger.warning(f"路徑查找檔案不存在: {lookup_file}")
            return
        
        try:
            import openpyxl
            workbook = openpyxl.load_workbook(lookup_file)
            
            # 假設資料在第一個工作表，格式為：課程名稱, 資源庫路徑, 資料夾路徑
            sheet = workbook.active
            
            for row in sheet.iter_rows(min_row=2, values_only=True):  # 跳過標題行
                if row[0]:  # 確保課程名稱不為空
                    course_name = str(row[0]).strip()
                    resource_path = str(row[1]).strip() if row[1] else ""
                    folder_path = str(row[2]).strip() if row[2] else ""
                    
                    self.path_lookup[course_name] = {
                        'folder_path': folder_path,
                        'resource_path': resource_path
                    }
            
            workbook.close()
            self.logger.info(f"成功載入 {len(self.path_lookup)} 筆路徑資料")
            
        except Exception as e:
            self.logger.error(f"載入路徑查找檔案失敗: {e}")
    
    def _get_path_info(self, sheet_name: str) -> Tuple[str, str]:
        """
        根據 Sheet 名稱查找路徑資訊，支援模糊查找
        
        Args:
            sheet_name: Sheet 名稱
            
        Returns:
            Tuple[str, str]: (資料夾路徑, 資源庫路徑)
        """
        # 首先嘗試精確匹配
        if sheet_name in self.path_lookup:
            info = self.path_lookup[sheet_name]
            self.logger.info(f"精確匹配成功: {sheet_name}")
            return info['folder_path'], info['resource_path']
        
        # 如果精確匹配失敗，進行模糊查找
        self.logger.info(f"精確匹配失敗，開始模糊查找: {sheet_name}")
        
        for course_name, info in self.path_lookup.items():
            # 檢查課程名稱是否包含 sheet_name
            if sheet_name in course_name:
                self.logger.info(f"模糊查找成功: {sheet_name} -> {course_name}")
                self.stats['fuzzy_matches'] += 1
                return info['folder_path'], info['resource_path']
        
        # 如果模糊查找也失敗
        self.logger.warning(f"模糊查找失敗，未找到匹配的課程: {sheet_name}")
        return "#未找到", "#未找到"
    
    def _create_output_directory(self):
        """建立輸出目錄"""
        try:
            self.output_dir.mkdir(parents=True, exist_ok=True)
            self.logger.info(f"輸出目錄已準備: {self.output_dir}")
        except Exception as e:
            self.logger.error(f"無法建立輸出目錄: {e}")
            raise
    
    def _is_youtube_url(self, href: str) -> bool:
        """檢查是否為 YouTube 連結"""
        if not href:
            return False
        return 'youtube.com' in href.lower() or 'youtu.be' in href.lower()
    
    def _is_video_file(self, path: str) -> bool:
        """檢查檔案是否為影片檔案"""
        if not path:
            return False
        
        # 移除查詢參數和錨點
        clean_path = path.split('?')[0].split('#')[0]
        extension = Path(clean_path).suffix.lower().lstrip('.')
        return extension in self.video_extensions
    
    def _is_audio_file(self, path: str) -> bool:
        """檢查檔案是否為音訊檔案"""
        if not path:
            return False
        
        # 移除查詢參數和錨點
        clean_path = path.split('?')[0].split('#')[0]
        extension = Path(clean_path).suffix.lower().lstrip('.')
        return extension in self.audio_extensions
    
    def _is_media_file(self, path: str) -> bool:
        """檢查檔案是否為影音檔案"""
        return self._is_video_file(path) or self._is_audio_file(path)
    
    def _is_image_file(self, path: str) -> bool:
        """檢查檔案是否為圖片檔案"""
        if not path:
            return False
        
        # 移除查詢參數和錨點
        clean_path = path.split('?')[0].split('#')[0]
        extension = Path(clean_path).suffix.lower().lstrip('.')
        return extension in self.image_extensions
    
    def _is_pdf_file(self, path: str) -> bool:
        """檢查檔案是否為 PDF 檔案"""
        if not path:
            return False
        
        # 移除查詢參數和錨點
        clean_path = path.split('?')[0].split('#')[0]
        extension = Path(clean_path).suffix.lower().lstrip('.')
        return extension in self.pdf_extensions
    
    def _is_html_file(self, path: str) -> bool:
        """檢查檔案是否為 HTML 檔案"""
        if not path:
            return False
        
        # 移除查詢參數和錨點
        clean_path = path.split('?')[0].split('#')[0]
        return clean_path.lower().endswith(('.html', '.htm'))
    
    def _determine_content_type(self, href: str, has_media_files: bool = False) -> str:
        """
        判斷內容類型 - 簡化版本
        
        Args:
            href: 檔案路徑或連結
            has_media_files: 是否包含影音檔案
            
        Returns:
            str: 內容類型
        """
        if not href:
            return "其他"
        
        # 移除註釋部分進行判斷
        clean_href = href.split('#')[0].strip()
        
        # 優先檢查 HTML 檔案 - 直接顯示為線上連結
        if self._is_html_file(clean_href):
            return "線上連結"
        
        # 如果有影音檔案，優先判斷為影音教材
        if has_media_files:
            return "影音教材"
        
        # YouTube 連結
        if self._is_youtube_url(clean_href):
            return "影音教材_影音連結"
        
        # 本地影片檔案
        if self._is_video_file(clean_href):
            return "影音教材_影片"
        
        # 本地音訊檔案
        if self._is_audio_file(clean_href):
            return "影音教材_音訊"
        
        # PDF 檔案
        if self._is_pdf_file(clean_href):
            return "參考檔案_PDF"
        
        # 圖片檔案
        if self._is_image_file(clean_href):
            return "參考檔案_圖片"
        
        return "其他"
    
    def _get_max_level(self, data: List[Dict[str, Any]]) -> int:
        """
        計算最大層級數
        
        Args:
            data: JSON 資料
            
        Returns:
            int: 最大層級數
        """
        max_level = 1  # 組織層級為1
        
        def _check_item_level(item: Dict[str, Any], current_level: int) -> int:
            max_in_branch = current_level
            if 'items' in item:
                for sub_item in item['items']:
                    sub_max = _check_item_level(sub_item, current_level + 1)
                    max_in_branch = max(max_in_branch, sub_max)
            return max_in_branch
        
        for org in data:
            if 'items' in org:
                for item in org['items']:
                    item_max = _check_item_level(item, 2)  # 組織下的項目從層級2開始
                    max_level = max(max_level, item_max)
        
        return max_level
    
    def _get_level_headers(self, max_level: int) -> List[str]:
        """
        根據層級數生成層級標題
        
        Args:
            max_level: 最大層級數
            
        Returns:
            List[str]: 層級標題列表
        """
        headers = []
        
        if max_level >= 1:
            headers.append("課程名稱")
        if max_level >= 2:
            headers.append("章節")
        
        # 從第3層開始是單元1, 單元2, ...
        for i in range(3, max_level):
            headers.append(f"單元{i-2}")
        
        if max_level >= 3:
            headers.append("學習活動")
        
        return headers
    
    def _extract_sheet_name(self, json_filename: str) -> str:
        """
        從 JSON 檔名提取 Sheet 名稱
        
        Args:
            json_filename: JSON 檔案名稱
            
        Returns:
            str: Sheet 名稱（第一個底線前的部分）
        """
        # 移除 .json 副檔名
        name_without_ext = json_filename.replace('.json', '')
        
        # 找到第一個底線的位置
        first_underscore = name_without_ext.find('_')
        
        if first_underscore != -1:
            return name_without_ext[:first_underscore]
        else:
            return name_without_ext  # 如果沒有底線，返回整個名稱
    
    def _has_href(self, item: Dict[str, Any]) -> bool:
        """
        檢查項目是否有 href
        
        Args:
            item: 項目資料
            
        Returns:
            bool: 是否有 href
        """
        return bool(item.get('href'))
    
    def _process_item(self, item: Dict[str, Any], level: int, max_level: int) -> List[Tuple[int, str, str, str, bool, str]]:
        """
        遞迴處理項目，生成 Excel 行資料
        
        Args:
            item: 項目資料
            level: 層級
            max_level: 最大層級數
            
        Returns:
            List[Tuple[int, str, str, str, bool, str]]: (層級, 標題, 類型, 路徑, 是否為學習活動, 學習活動標記) 的列表
        """
        rows = []
        
        title = item.get('title', '')
        href = item.get('href', '')
        has_media_files = bool(item.get('media_files'))
        has_href = self._has_href(item)
        has_children = bool(item.get('items'))
        
        # 判斷內容類型
        content_type = self._determine_content_type(href, has_media_files) if href else ""
        
        if has_href and has_children:
            # 有 href 且有子項目：保留在原層級，學習活動欄位標記 "#Yes"
            rows.append((level, title, content_type, href, True, "#Yes"))
            self.stats['items_processed'] += 1
        elif has_href and not has_children:
            # 有 href 但沒有子項目：直接放在學習活動欄位
            rows.append((max_level, title, content_type, href, True, ""))
            self.stats['items_processed'] += 1
        else:
            # 沒有 href：視為分類標題，保留在原層級
            rows.append((level, title, "", "", False, ""))
        
        # 處理子項目
        if 'items' in item:
            for sub_item in item['items']:
                sub_rows = self._process_item(sub_item, level + 1, max_level)
                rows.extend(sub_rows)
        
        return rows
    
    def _create_sheet_header(self, worksheet, sheet_name: str, json_filename: str, max_level: int):
        """
        建立 Sheet 表頭
        
        Args:
            worksheet: 工作表物件
            sheet_name: Sheet 名稱
            json_filename: JSON 檔案名稱
            max_level: 最大層級數
        """
        # 查找路徑資訊
        folder_path, resource_path = self._get_path_info(sheet_name)
        
        # 設定表頭資訊
        worksheet['A1'] = f"名稱：{sheet_name}"
        worksheet['A2'] = f"資源庫路徑：{resource_path}"
        worksheet['A3'] = f"資料夾路徑：{folder_path}"
        
        # 計算類型和路徑的欄位位置（層級數 + 1個空格 + 類型/路徑）
        type_col = max_level + 2
        path_col = max_level + 3
        
        # 設定欄位標題（第5行）
        worksheet[f'{get_column_letter(1)}5'] = "層級"
        worksheet[f'{get_column_letter(type_col)}5'] = "類型"  
        worksheet[f'{get_column_letter(path_col)}5'] = "路徑"
        
        # 設定層級標題（第6行）
        level_headers = self._get_level_headers(max_level)
        for i, header in enumerate(level_headers):
            col_letter = get_column_letter(i + 1)
            worksheet[f'{col_letter}6'] = header
        
        # 設定字體樣式
        header_font = Font(bold=True)
        for row in [1, 2, 3, 5, 6]:
            for col_idx in range(1, max(path_col + 1, 7)):  # 至少到G欄
                col_letter = get_column_letter(col_idx)
                cell = worksheet[f'{col_letter}{row}']
                if cell.value:
                    cell.font = header_font
    
    def _write_data_to_sheet(self, worksheet, rows: List[Tuple[int, str, str, str, bool, str]], max_level: int):
        """
        將資料寫入工作表
        
        Args:
            worksheet: 工作表物件
            rows: 資料行列表
            max_level: 最大層級數
        """
        start_row = 7  # 從第7行開始寫入資料
        
        # 計算類型和路徑的欄位位置
        type_col = max_level + 2
        path_col = max_level + 3
        
        for i, (level, title, content_type, path, is_activity, activity_mark) in enumerate(rows):
            current_row = start_row + i
            
            # 根據層級寫入對應的欄位
            if level >= 1 and level <= max_level:
                col_letter = get_column_letter(level)
                worksheet[f'{col_letter}{current_row}'] = title
            
            # 如果有學習活動標記，在學習活動欄位寫入標記
            if activity_mark:
                worksheet[f'{get_column_letter(max_level)}{current_row}'] = activity_mark
            
            # 有學習活動的項目寫入類型和路徑
            if is_activity:
                if content_type:
                    worksheet[f'{get_column_letter(type_col)}{current_row}'] = content_type
                if path:
                    worksheet[f'{get_column_letter(path_col)}{current_row}'] = path
    
    def _process_json_file(self, json_file: Path, workbook) -> bool:
        """
        處理單個 JSON 檔案
        
        Args:
            json_file: JSON 檔案路徑
            workbook: Excel 工作簿物件
            
        Returns:
            bool: 處理是否成功
        """
        try:
            # 讀取 JSON 檔案
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # 提取 Sheet 名稱
            sheet_name = self._extract_sheet_name(json_file.name)
            
            # 計算最大層級數
            max_level = self._get_max_level(data)
            
            # 建立工作表
            worksheet = workbook.create_sheet(title=sheet_name)
            
            # 建立表頭
            json_filename = json_file.name
            self._create_sheet_header(worksheet, sheet_name, json_filename, max_level)
            
            # 處理所有組織
            all_rows = []
            for org in data:
                org_title = org.get('title', '')
                if org_title:
                    all_rows.append((1, org_title, "", "", False, ""))  # 組織層級為1
                
                # 處理組織下的項目
                if 'items' in org:
                    for item in org['items']:
                        item_rows = self._process_item(item, 2, max_level)  # 組織下的項目從層級2開始
                        all_rows.extend(item_rows)
            
            # 寫入資料到工作表
            self._write_data_to_sheet(worksheet, all_rows, max_level)
            
            # 調整欄寬（根據實際的欄位數）
            total_cols = max_level + 3  # 層級數 + 空格 + 類型 + 路徑
            for col_idx in range(1, total_cols + 1):
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 20
            
            self.stats['sheets_created'] += 1
            self.logger.info(f"成功處理 JSON 檔案: {json_file.name} -> Sheet: {sheet_name}")
            return True
            
        except Exception as e:
            self.stats['errors'] += 1
            self.logger.error(f"處理 JSON 檔案失敗 {json_file.name}: {e}")
            return False
    
    def generate_excel(self) -> bool:
        """
        生成 Excel 檔案
        
        Returns:
            bool: 生成是否成功
        """
        try:
            # 檢查 JSON 目錄是否存在
            if not self.json_dir.exists():
                self.logger.error(f"JSON 目錄不存在: {self.json_dir}")
                return False
            
            # 建立輸出目錄
            self._create_output_directory()
            
            # 生成統一的時間戳
            self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # 找到所有 JSON 檔案，略過 path_mappings.json
            json_files = [f for f in self.json_dir.glob("*.json") if f.name != "path_mappings.json"]
            if not json_files:
                self.logger.warning(f"在目錄 {self.json_dir} 中未找到可處理的 JSON 檔案")
                return False
            
            # 建立 Excel 工作簿
            workbook = openpyxl.Workbook()
            
            # 移除預設的工作表
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            
            # 處理每個 JSON 檔案
            print(f"🔍 找到 {len(json_files)} 個 JSON 檔案")
            
            for json_file in sorted(json_files):
                print(f"正在處理: {json_file.name}")
                success = self._process_json_file(json_file, workbook)
                if success:
                    self.stats['json_files_processed'] += 1
            
            # 生成輸出檔名
            output_filename = f"course_structures_{self.timestamp}.xlsx"
            self.output_file = self.output_dir / output_filename
            
            # 儲存 Excel 檔案
            workbook.save(self.output_file)
            
            self.logger.info(f"Excel 檔案已儲存: {self.output_file}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"生成 Excel 檔案時發生錯誤: {e}")
            return False
    
    def print_summary(self):
        """輸出處理摘要"""
        print("\n" + "="*50)
        print("📊 Excel 生成作業完成")
        print("="*50)
        print(f"處理 JSON 檔案數: {self.stats['json_files_processed']}")
        print(f"建立 Sheet 數: {self.stats['sheets_created']}")
        print(f"處理項目數: {self.stats['items_processed']}")
        if self.stats['fuzzy_matches'] > 0:
            print(f"模糊查找成功數: {self.stats['fuzzy_matches']}")
        if self.stats['errors'] > 0:
            print(f"錯誤數: {self.stats['errors']}")
        
        if self.output_file:
            print(f"\n📄 課程結構檔案: {self.output_file.name}")
            print(f"檔案位置: {self.output_file.parent.absolute()}")


def main():
    """主函數"""
    print("🚀 Excel 課程結構生成器")
    print("="*30)
    
    # 檢查依賴
    try:
        import openpyxl
    except ImportError:
        print("❌ 缺少必要的依賴套件：openpyxl")
        print("請執行：pip install openpyxl")
        return False
    
    # 取得用戶輸入
    while True:
        print("請輸入 JSON 檔案資料夾名稱 (輸入 '0' 使用預設: 4_manifest_structures): ", end="", flush=True)
        json_folder = input().strip()
        if not json_folder:
            print("⚠️ 請輸入有效值，或輸入 '0' 使用預設值")
            continue
        if json_folder == '0':
            json_folder = "4_manifest_structures"
        
        json_path = Path(json_folder)
        if json_path.exists():
            break
        else:
            print(f"❌ 資料夾 '{json_folder}' 不存在，請重新輸入")
    
    # 建立生成器並執行
    generator = ExcelGenerator(json_folder)
    success = generator.generate_excel()
    
    # 輸出摘要
    generator.print_summary()
    
    if success and generator.stats['sheets_created'] > 0:
        print(f"\n🎉 生成完成！")
        print(f"📁 課程結構檔案：{generator.output_file.name}")
        print(f"📂 檔案位置：{generator.output_dir.absolute()}")
    elif success:
        print(f"\n⚠️  作業完成，但沒有建立任何 Sheet")
    else:
        print(f"\n❌ 生成過程中發生錯誤，請檢查日誌檔案")
    
    return success


if __name__ == "__main__":
    main()